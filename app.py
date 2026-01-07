import streamlit as st
import pandas as pd
import random
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from collections import defaultdict

# ==========================================
# 1. 페이지 설정 및 디자인
# ==========================================
st.set_page_config(page_title="근무표 자동화 시스템", layout="wide")

st.markdown("""
    <style>
        .stTable { border: 1px solid #333; font-size: 14px; }
        th { background-color: #F2F2F2 !important; color: black !important; font-weight: bold !important; text-align: center !important; border: 1px solid #333 !important; }
        td { border: 1px solid #333 !important; text-align: center !important; }
        .stButton>button { width: 100%; }
    </style>
""", unsafe_allow_html=True)

# 설정값
LOCATIONS_CONFIG = {
    "인천": {"생활관1": 2, "생활관2": 2, "생활관3": 2, "상황실1": 3, "도서관1": 2},
    "경기": {"생활관1": 2, "생활관2": 2, "상황실2": 3, "도서관2": 2}
}
HOLIDAYS = ['2025-10-03', '2025-10-06', '2025-10-09']

# ==========================================
# 2. 핵심 로직 (코랩 배정 엔진)
# ==========================================

def get_korean_weekday(date_obj):
    return ['월', '화', '수', '목', '금', '토', '일'][date_obj.weekday()]

def generate_schedule_logic(df_staff, start_dt, end_dt):
    df_staff['이름'] = df_staff['이름'].astype(str).str.strip()
    work_counts = {name: 0 for name in df_staff['이름'].unique()}
    schedule_results = []
    
    fixed_assignments = defaultdict(list)
    for _, row in df_staff.iterrows():
        if pd.notna(row.get('고정근무일자')):
            raw_dates = str(row['고정근무일자']).split(',')
            raw_locs = str(row['고정근무지']).split(',') if pd.notna(row.get('고정근무지')) else []
            for i, d_str in enumerate(raw_dates):
                try:
                    clean_date = datetime.strptime(d_str.strip(), '%Y-%m-%d').strftime('%Y-%m-%d')
                    loc_target = raw_locs[i].strip() if i < len(raw_locs) else (raw_locs[0].strip() if raw_locs else "미지정")
                    fixed_assignments[clean_date].append((row['이름'], loc_target, row['캠퍼스']))
                    work_counts[row['이름']] += 1
                except: continue

    date_range = []
    curr = start_dt
    while curr <= end_dt:
        if curr.weekday() < 5 and curr.strftime("%Y-%m-%d") not in HOLIDAYS:
            date_range.append(curr)
        curr += timedelta(days=1)

    for date in date_range:
        date_str = date.strftime("%Y-%m-%d")
        today_assigned = []
        if date_str in fixed_assignments:
            for name, loc, campus in fixed_assignments[date_str]:
                schedule_results.append({"날짜": date_str, "캠퍼스": campus, "근무지": loc, "직원": name, "유형": "고정"})
                today_assigned.append(name)

        for campus, locs in LOCATIONS_CONFIG.items():
            for loc_name, total_required in locs.items():
                already_filled = len([s for s in schedule_results if s['날짜'] == date_str and s['캠퍼스'] == campus and s['근무지'] == loc_name])
                needed = total_required - already_filled
                if needed <= 0: continue
                
                possible_staff = df_staff[((df_staff['캠퍼스'] == campus) | (df_staff['캠퍼스'] == "모두")) & (~df_staff['이름'].isin(today_assigned))]
                final_candidates = []
                for _, s_row in possible_staff.iterrows():
                    dept = str(s_row['소속'])
                    is_excluded = any(key in dept and key in loc_name for key in ['생활관', '상황실', '도서관'])
                    if not is_excluded: final_candidates.append(s_row['이름'])

                random.shuffle(final_candidates)
                final_candidates.sort(key=lambda x: work_counts[x])
                assigned_now = final_candidates[:needed]
                for person in assigned_now:
                    schedule_results.append({"날짜": date_str, "캠퍼스": campus, "근무지": loc_name, "직원": person, "유형": "일반"})
                    work_counts[person] += 1
                    today_assigned.append(person)

    return pd.DataFrame(schedule_results), work_counts

# ==========================================
# 3. 엑셀 생성 (코랩 스타일 복원)
# ==========================================

def make_final_excel_blob(df, stats):
    output = BytesIO()
    wb = Workbook()
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    import streamlit as st
import pandas as pd
import random
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from collections import defaultdict

# ==========================================
# 1. 초기 설정 및 디자인
# ==========================================
st.set_page_config(page_title="근무표 통합 시스템", layout="wide", initial_sidebar_state="collapsed")

# 세션 데이터 초기화 (앱이 실행되는 동안 데이터 유지)
if 'df' not in st.session_state: st.session_state.df = None
if 'stats' not in st.session_state: st.session_state.stats = {}
if 'last_updated' not in st.session_state: st.session_state.last_updated = "기록 없음"

# CSS: 가독성 및 디자인 최적화
st.markdown("""
    <style>
        .main-title { font-size: 2.3rem; font-weight: bold; color: #1E3A8A; text-align: center; margin-bottom: 0px; }
        .update-box { background-color: #f0f2f6; padding: 10px; border-radius: 10px; text-align: center; margin-bottom: 25px; font-size: 0.9rem; color: #555; }
        .stTable { border: 1px solid #ddd; }
        th { background-color: #1E3A8A !important; color: white !important; text-align: center !important; }
        td { text-align: center !important; }
    </style>
""", unsafe_allow_html=True)

# 상수 설정
LOCATIONS_CONFIG = {
    "인천": {"생활관1": 2, "생활관2": 2, "생활관3": 2, "상황실1": 3, "도서관1": 2},
    "경기": {"생활관1": 2, "생활관2": 2, "상황실2": 3, "도서관2": 2}
}
HOLIDAYS = ['2025-10-03', '2025-10-06', '2025-10-09']

# ==========================================
# 2. 유틸리티 함수 (요일, 엑셀생성, 로직)
# ==========================================

def get_korean_weekday(date_obj):
    return ['월', '화', '수', '목', '금', '토', '일'][date_obj.weekday()]

def update_timestamp():
    st.session_state.last_updated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def make_final_excel_blob(df, stats):
    """요청하신 3개 시트 구성 엑셀 생성"""
    output = BytesIO()
    wb = Workbook()
    
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 시트 1: 주간근무표 (시각화 세로 나열)
    ws1 = wb.active
    ws1.title = "주간근무표"
    dates = sorted(df['날짜'].unique())
    curr_r = 1
    
    for d_str in dates:
        dt_obj = datetime.strptime(d_str, '%Y-%m-%d') if isinstance(d_str, str) else d_str
        d_label = d_str if isinstance(d_str, str) else d_str.strftime('%Y-%m-%d')
        
        ws1.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=6)
        cell = ws1.cell(row=curr_r, column=1, value=f"{d_label}({get_korean_weekday(dt_obj)}) 근무표")
        cell.alignment = center; cell.fill = header_fill; cell.font = Font(bold=True)
        curr_r += 1
        
        headers = ["캠퍼스", "도서관", "상황실", "생활관1", "생활관2", "생활관3"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=curr_r, column=c_idx, value=h)
            cell.alignment = center; cell.border = border; cell.fill = header_fill
        curr_r += 1
        
        for cp in ["인천", "경기"]:
            ws1.cell(row=curr_r, column=1, value=cp).border = border
            for c_idx, loc_b in enumerate(["도서관", "상황실", "생활관1", "생활관2", "생활관3"], 2):
                loc_f = loc_b if "생활관" in loc_b else loc_b + ("1" if cp=="인천" else "2")
                names = df[(df['날짜']==d_label) & (df['캠퍼스']==cp) & (df['근무지']==loc_f)]['직원'].tolist()
                cell = ws1.cell(row=curr_r, column=c_idx, value=", ".join(names))
                cell.border = border; cell.alignment = center
            curr_r += 1
        curr_r += 1 # 공백

    # 시트 2: 전체데이터
    ws2 = wb.create_sheet("전체데이터")
    for r_idx, row in enumerate([df.columns.tolist()] + df.values.tolist(), 1):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=str(val)).border = border

    # 시트 3: 근무통계
    ws3 = wb.create_sheet("근무통계")
    ws3.append(["직원 이름", "횟수"])
    for name, count in stats.items():
        ws3.append([name, count])

    wb.save(output)
    return output.getvalue()

def generate_schedule_logic(df_staff, start_dt, end_dt):
    df_staff['이름'] = df_staff['이름'].astype(str).str.strip()
    work_counts = {name: 0 for name in df_staff['이름'].unique()}
    schedule_results = []
    
    fixed_assignments = defaultdict(list)
    for _, row in df_staff.iterrows():
        if pd.notna(row.get('고정근무일자')):
            raw_dates = str(row['고정근무일자']).split(',')
            raw_locs = str(row['고정근무지']).split(',') if pd.notna(row.get('고정근무지')) else []
            for i, d_str in enumerate(raw_dates):
                try:
                    clean_date = datetime.strptime(d_str.strip(), '%Y-%m-%d').strftime('%Y-%m-%d')
                    loc_target = raw_locs[i].strip() if i < len(raw_locs) else (raw_locs[0].strip() if raw_locs else "미지정")
                    fixed_assignments[clean_date].append((row['이름'], loc_target, row['캠퍼스']))
                    work_counts[row['이름']] += 1
                except: continue

    date_range = []
    curr = start_dt
    while curr <= end_dt:
        if curr.weekday() < 5 and curr.strftime("%Y-%m-%d") not in HOLIDAYS:
            date_range.append(curr)
        curr += timedelta(days=1)

    for date in date_range:
        date_str = date.strftime("%Y-%m-%d")
        today_assigned = []
        if date_str in fixed_assignments:
            for name, loc, campus in fixed_assignments[date_str]:
                schedule_results.append({"날짜": date_str, "캠퍼스": campus, "근무지": loc, "직원": name, "유형": "고정"})
                today_assigned.append(name)

        for campus, locs in LOCATIONS_CONFIG.items():
            for loc_name, total_required in locs.items():
                already_filled = len([s for s in schedule_results if s['날짜'] == date_str and s['캠퍼스'] == campus and s['근무지'] == loc_name])
                needed = total_required - already_filled
                if needed <= 0: continue
                
                possible_staff = df_staff[((df_staff['캠퍼스'] == campus) | (df_staff['캠퍼스'] == "모두")) & (~df_staff['이름'].isin(today_assigned))]
                final_candidates = []
                for _, s_row in possible_staff.iterrows():
                    dept = str(s_row['소속'])
                    is_excluded = any(key in dept and key in loc_name for key in ['생활관', '상황실', '도서관'])
                    if not is_excluded: final_candidates.append(s_row['이름'])

                random.shuffle(final_candidates)
                final_candidates.sort(key=lambda x: work_counts[x])
                assigned_now = final_candidates[:needed]
                for person in assigned_now:
                    schedule_results.append({"날짜": date_str, "캠퍼스": campus, "근무지": loc_name, "직원": person, "유형": "일반"})
                    work_counts[person] += 1
                    today_assigned.append(person)

    return pd.DataFrame(schedule_results), work_counts

# ==========================================
# 3. 사이드바 (관리자 메뉴)
# ==========================================
with st.sidebar:
    st.title("⚙️ 관리자 설정")
    admin_pw = st.text_input("관리자 비밀번호", type="password")
    
    if admin_pw == "1234":
        st.success("로그인 성공")
        st.divider()
        st.subheader("1️⃣ 새 근무표 생성")
        file = st.file_uploader("명단 업로드(.xlsx)", type=['xlsx'])
        s_date = st.date_input("시작일", datetime.today())
        e_date = st.date_input("종료일", datetime.today() + timedelta(days=14))
        
        if st.button("🚀 근무표 자동 생성"):
            if file:
                input_df = pd.read_excel(file)
                res_df, res_stats = generate_schedule_logic(input_df, s_date, e_date)
                st.session_state.df = res_df
                st.session_state.stats = res_stats
                update_timestamp()
                st.rerun()
        
        st.divider()
        if st.session_state.df is not None:
            st.subheader("2️⃣ 근무자 수동 교체")
            df_edit = st.session_state.df
            idx1 = st.selectbox("대상 1", df_edit.index, format_func=lambda x: f"{df_edit.loc[x,'날짜']} | {df_edit.loc[x,'직원']}")
            idx2 = st.selectbox("대상 2", df_edit.index, format_func=lambda x: f"{df_edit.loc[x,'날짜']} | {df_edit.loc[x,'직원']}")
            if st.button("🪄 교체 확정"):
                df_edit.at[idx1, '직원'], df_edit.at[idx2, '직원'] = df_edit.at[idx2, '직원'], df_edit.at[idx1, '직원']
                st.session_state.df = df_edit
                st.session_state.stats = df_edit['직원'].value_counts().to_dict()
                update_timestamp()
                st.success("교체 완료!")
                st.rerun()
    else:
        st.info("비밀번호를 입력하면 관리자 도구가 나타납니다.")

# ==========================================
# 4. 메인 화면 (직원 배포용 뷰)
# ==========================================
st.markdown('<div class="main-title">📅 근무 배정 게시판</div>', unsafe_allow_html=True)
st.markdown(f'<div class="update-box"><strong>최종 업데이트:</strong> {st.session_state.last_updated}</div>', unsafe_allow_html=True)

if st.session_state.df is not None:
    # [상단 도구] 다운로드 및 검색
    c1, c2 = st.columns([1, 1])
    with c1:
        excel_data = make_final_excel_blob(st.session_state.df, st.session_state.stats)
        st.download_button("📥 엑셀 다운로드 (전체 시트 포함)", excel_data, 
                           file_name=f"근무표_{datetime.now().strftime('%m%d')}.xlsx", use_container_width=True)
    with c2:
        search_name = st.text_input("🔍 내 근무 찾기 (이름 입력)", "")

    if search_name:
        search_res = st.session_state.df[st.session_state.df['직원'].str.contains(search_name)]
        st.markdown(f"**'{search_name}'** 님의 검색 결과입니다.")
        st.table(search_res)

    st.divider()

    # [메인 근무표] 주차별 세로 나열
    temp_df = st.session_state.df.copy()
    temp_df['날짜'] = pd.to_datetime(temp_df['날짜'])
    temp_df['주차'] = temp_df['날짜'].dt.isocalendar().week
    
    for i, week in enumerate(sorted(temp_df['주차'].unique())):
        st.subheader(f"📍 {i+1}주차 근무 현황")
        week_data = temp_df[temp_df['주차'] == week]
        pivot = week_data.pivot_table(
            index=['캠퍼스', '근무지'], columns='날짜', values='직원', aggfunc=lambda x: ", ".join(x)
        ).fillna("-")
        pivot.columns = [f"{d.strftime('%m-%d')}({get_korean_weekday(d)})" for d in pivot.columns]
        st.table(pivot)
else:
    st.info("게시된 근무표가 없습니다. 관리자가 파일을 업로드해야 합니다.")